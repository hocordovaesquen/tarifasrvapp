import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook

# --------------------------------------------------
# Utilidades de carga desde el Excel
# --------------------------------------------------

def load_bbdd_from_bytes(data: bytes, sheet_name: str = "A.3 BBDD Neg") -> pd.DataFrame:
    """
    Carga y limpia la hoja A.3 BBDD Neg usando la fila 6 (header=5) como cabecera cruda
    y la fila 1 de esa lectura como "nombres reales" de las columnas.
    """
    raw = pd.read_excel(BytesIO(data), sheet_name=sheet_name, header=5)
    header = raw.iloc[0]

    cols = [str(header[c]) if not pd.isna(header[c]) else str(c) for c in raw.columns]
    df = raw.iloc[1:].copy()
    df.columns = cols

    num_cols = [
        "Monto Local",
        "Monto USD",
        "Monto DMA Local",
        "Monto DMA USD",
        "Cobro Acceso",
        "Cobro Transacción",
        "Cobro Transacción2",
        "Cobro Perfiles",
        "Total ingresos",
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    return df


def load_access_tramos_from_bytes(data: bytes) -> dict:
    """
    Lee los tramos de ACCESO OP DIRECTO INSTITUCIONAL desde:
    Hoja '1. Parametros' / '1. Parámetros' – bloque que arranca en T89.

    Estructura resultado:
    {
        "Colombia": [(min, max, fija_usd), ...],
        "Perú":     [(min, max, fija_usd), ...],
        "Chile":    [(min, max, fija_usd), ...],
    }
    """
    # Abrimos el workbook
    wb = load_workbook(BytesIO(data), data_only=True)

    # Intentamos ambas variantes de nombre de hoja por si el archivo tiene tilde o no
    if "1. Parametros" in wb.sheetnames:
        ws = wb["1. Parametros"]
    elif "1. Parámetros" in wb.sheetnames:
        ws = wb["1. Parámetros"]
    else:
        raise ValueError("No se encontró la hoja '1. Parametros' ni '1. Parámetros' en el Excel.")

    # Filas donde están los Tramo 1, 2 y 3
    rows = [91, 92, 93]

    # Mapeo de columnas: Min, Max, Fija(USD) por país
    coords = {
        "Colombia": ("T", "U", "W"),
        "Perú":     ("X", "Y", "AA"),
        "Chile":    ("AB", "AC", "AE"),
    }

    res = {}
    for country, (col_min, col_max, col_fija) in coords.items():
        tramos = []
        for r in rows:
            minv = ws[f"{col_min}{r}"].value
            maxv = ws[f"{col_max}{r}"].value
            fijav = ws[f"{col_fija}{r}"].value
            if minv is None or maxv is None or fijav is None:
                continue
            tramos.append((float(minv), float(maxv), float(fijav)))
        res[country] = tramos

    return res


def inicializar_tramos_tx(df: pd.DataFrame) -> pd.DataFrame:
    """
    Genera una propuesta de tramos inicial para la parte TRANSACCIONAL
    usando cuantiles de la base (Monto USD + Monto DMA USD).
    """
    base = df["Monto USD"] + df.get("Monto DMA USD", 0)
    base = base.fillna(0)

    if base.sum() == 0:
        tramos = pd.DataFrame(
            {
                "Desde": [0.0],
                "Hasta": [np.nan],
                "Fijo_USD": [500.0],
                "Variable_bps": [10.0],
            }
        )
        return tramos

    q1 = float(base.quantile(0.25))
    q2 = float(base.quantile(0.50))
    q3 = float(base.quantile(0.75))

    tramos = pd.DataFrame(
        {
            "Desde": [0.0, q1, q2, q3],
            "Hasta": [q1, q2, q3, np.nan],
            "Fijo_USD": [500.0, 500.0, 500.0, 500.0],
            "Variable_bps": [5.0, 8.0, 10.0, 12.0],
        }
    )
    return tramos


# --------------------------------------------------
# Cálculo de ingresos con parte TRANSACCIONAL + ACCESO
# --------------------------------------------------

def compute_new_fees(df: pd.DataFrame,
                     tramos_tx: pd.DataFrame,
                     access_tramos: dict,
                     adoption: float = 1.0) -> pd.DataFrame:
    """
    Calcula ingresos nuevos combinando:
    - Parte TRANSACCIONAL (editable en el dashboard, tramos_tx)
    - Parte de ACCESO (desde hoja 1. Parametros, bloque T89+)

    adoption: número entre 0 y 1 que representa % de implementación.
    """
    df = df.copy()

    if "Monto USD" not in df.columns:
        st.error("No se encontró la columna 'Monto USD' en la BBDD.")
        return df

    base = df["Monto USD"].astype(float) + df.get("Monto DMA USD", 0).astype(float)

    # ---------------------------
    # 1. NUEVA PARTE TRANSACCIONAL
    # ---------------------------
    new_tx_full = pd.Series(0.0, index=df.index)

    for _, row in tramos_tx.iterrows():
        try:
            desde = float(row["Desde"])
        except Exception:
            desde = 0.0

        hasta = row.get("Hasta", np.nan)
        try:
            fijo = float(row["Fijo_USD"])
        except Exception:
            fijo = 0.0

        try:
            bps = float(row["Variable_bps"])
        except Exception:
            bps = 0.0

        if pd.isna(hasta):
            mask = base >= desde
        else:
            mask = (base >= desde) & (base < float(hasta))

        var_component = base[mask] * (bps / 10000.0)  # bps a %
        new_tx_full.loc[mask] = fijo + var_component

    # ---------------------------
    # 2. NUEVA PARTE DE ACCESO (FIJA POR TRAMO)
    # ---------------------------
    new_access_full = pd.Series(0.0, index=df.index)

    if "Pais" not in df.columns:
        st.warning("No se encontró la columna 'Pais' para aplicar tramos de acceso por país.")
    else:
        for country, tr_list in access_tramos.items():
            mask_country = df["Pais"] == country
            if not mask_country.any():
                continue

            base_country = base[mask_country]
            access_vals = pd.Series(0.0, index=base_country.index)

            for (mn, mx, fija) in tr_list:
                # mx puede ser muy grande (9.99e20) para el último tramo
                if mx is None:
                    mask_tr = base_country >= mn
                else:
                    mask_tr = (base_country >= mn) & (base_country <= mx)
                access_vals.loc[mask_tr] = fija

            new_access_full.loc[mask_country] = access_vals

    # ---------------------------
    # 3. RECONSTRUCCIÓN DE INGRESOS
    # ---------------------------
    if "Total ingresos" not in df.columns:
        st.error("No se encontró la columna 'Total ingresos' en la BBDD.")
        return df

    actual_total = df["Total ingresos"].astype(float)

    current_tx = df.get("Cobro Transacción2", 0.0)
    if isinstance(current_tx, pd.Series):
        current_tx = current_tx.astype(float)

    current_access = df.get("Cobro Acceso", 0.0)
    if isinstance(current_access, pd.Series):
        current_access = current_access.astype(float)

    # Otros componentes (perfiles, etc.) se mantienen
    other_components = actual_total - current_tx - current_access

    # Escenario full: nuevo acceso + nueva transacción + otros
    new_total_full = other_components + new_tx_full + new_access_full

    # Escenario parcial según % implementación
    new_total_partial = actual_total + adoption * (new_total_full - actual_total)

    df["Ingreso_actual"] = actual_total
    df["Ingreso_nuevo_full"] = new_total_full
    df["Ingreso_nuevo_escenario"] = new_total_partial
    df["Nuevo_tx_full"] = new_tx_full
    df["Nuevo_acceso_full"] = new_access_full

    return df


# --------------------------------------------------
# UI principal Streamlit
# --------------------------------------------------

st.set_page_config(
    page_title="Simulador Estructura Tarifaria RV - nuam",
    layout="wide",
)

st.title("📊 Simulador de Estructura Tarifaria RV")
st.caption(
    "Modelo basado en tu Excel: recalcula parte transaccional y acceso (T89+) por tramos, "
    "y permite simular % de implementación."
)

st.sidebar.header("⚙️ Configuración")

uploaded_file = st.sidebar.file_uploader(
    "Sube el modelo RV (.xlsx)",
    type=["xlsx"],
    help="Usa el archivo '23102025 Modelamiento Estructura Tarifaria RV (1).xlsx'",
)

sheet_name = st.sidebar.text_input(
    "Nombre de hoja BBDD",
    value="A.3 BBDD Neg",
    help="Hoja base de datos de negociación (BBDD).",
)

adopcion_pct = st.sidebar.slider(
    "Porcentaje de implementación del nuevo tarifario",
    min_value=0,
    max_value=100,
    value=50,
    step=5,
)
adoption = adopcion_pct / 100.0

st.sidebar.markdown("---")
st.sidebar.write("Ajusta los tramos TRANSACCIONALES abajo. El acceso se lee de '1. Parametros' (T89+).")

if uploaded_file is None:
    st.info("📥 Sube el archivo Excel en la barra lateral para comenzar.")
    st.stop()

# --------------------------------------------------
# Carga de Excel completo una sola vez (bytes)
# --------------------------------------------------

uploaded_bytes = uploaded_file.getvalue()

with st.spinner("Cargando BBDD y parámetros de acceso..."):
    try:
        bbdd_df = load_bbdd_from_bytes(uploaded_bytes, sheet_name=sheet_name)
        access_tramos = load_access_tramos_from_bytes(uploaded_bytes)
    except Exception as e:
        st.error(f"Error cargando el archivo: {e}")
        st.stop()

st.success("BBDD y tramos de acceso cargados correctamente.")

# --------------------------------------------------
# Tramos TRANSACCIONALES editables
# --------------------------------------------------

if "tramos_tx_df" not in st.session_state:
    st.session_state["tramos_tx_df"] = inicializar_tramos_tx(bbdd_df)

st.subheader("🧩 Tramos TRANSACCIONALES propuestos")

st.markdown(
    """
Cada fila representa un **tramo de volumen** (en USD) para la parte de **transacción**:

- **Desde / Hasta**: límites del tramo en USD (si `Hasta` está vacío, aplica "en adelante").
- **Fijo_USD**: componente fijo del cobro transaccional.
- **Variable_bps**: componente variable en *basis points* (ej. 25 = 0.25%).

La parte de **acceso** se toma automáticamente de la hoja `1. Parametros` / `1. Parámetros`
(bloque que inicia en **T89**).
"""
)

edited_tramos = st.data_editor(
    st.session_state["tramos_tx_df"],
    num_rows="dynamic",
    use_container_width=True,
    key="editor_tramos_tx",
)

st.session_state["tramos_tx_df"] = edited_tramos

# --------------------------------------------------
# Cálculo de escenario
# --------------------------------------------------

st.subheader("🚀 Cálculo de escenario de ingresos (Transacción + Acceso)")

with st.spinner("Calculando ingresos actuales y proyectados..."):
    scenario_df = compute_new_fees(
        bbdd_df,
        st.session_state["tramos_tx_df"],
        access_tramos,
        adoption=adoption,
    )

# KPIs globales
total_actual = scenario_df["Ingreso_actual"].sum()
total_full = scenario_df["Ingreso_nuevo_full"].sum()
total_scenario = scenario_df["Ingreso_nuevo_escenario"].sum()

delta_full = total_full - total_actual
delta_pct_full = (delta_full / total_actual * 100) if total_actual != 0 else 0.0

delta_scenario = total_scenario - total_actual
delta_pct_scenario = (
    (delta_scenario / total_actual * 100) if total_actual != 0 else 0.0
)

col1, col2, col3 = st.columns(3)

col1.metric(
    "Ingreso actual (USD)",
    f"{total_actual:,.0f}",
)

col2.metric(
    f"Ingreso nuevo (100% implementación)",
    f"{total_full:,.0f}",
    f"{delta_full:,.0f} ({delta_pct_full:,.1f}%)",
)

col3.metric(
    f"Ingreso escenario ({adopcion_pct}%)",
    f"{total_scenario:,.0f}",
    f"{delta_scenario:,.0f} ({delta_pct_scenario:,.1f}%)",
)

st.markdown("---")

# ----------------------
